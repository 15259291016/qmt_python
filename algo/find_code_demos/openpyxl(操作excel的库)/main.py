import pandas as pd
from datetime import datetime, timedelta, time

def get_trade_date(day: int = 0):
    """
    获取交易日
    """
    days_to_subtract = 5
    skipped_days = 0

    while skipped_days < days_to_subtract:
        current_date = datetime.now()
        current_date = current_date - timedelta(days=day)
        if current_date.weekday() < 5:  # 周一到周五为0-4
            skipped_days += 1
        else:
            day += 1
    return current_date.strftime('%Y%m%d')

# data = {'日期': [get_trade_date(i) for i in range(0,4)],
#         'Age': [28, 34, 29, 32],
#         'Job': ['Engineer', 'Doctor', 'Architect', 'Teacher']}

# df = pd.DataFrame(data)


# # 将DataFrame写入Excel文件
# df.to_excel('output.xlsx', sheet_name='Sheet1', index=False)

from openpyxl import load_workbook
# 读
wb = load_workbook(filename='output.xlsx')
sheet = wb['Sheet1']
# 读取特定单元格的值
print(sheet['A1'].value)
# 或者
print(sheet.cell(row=1, column=1).value)

# 遍历行
for row in sheet.iter_rows(values_only=True):
    print(row)
    
    
# 写
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Border, Side, Alignment

# 创建一个新的工作簿
wb = Workbook()
sheet = wb.active

# 设置单元格的字体样式
font = Font(name='Calibri', size=12, bold=True, italic=True, color='FF0000')
sheet['A1'].font = font
sheet['A1'] = 'Hello, World!'
fill = PatternFill(fill_type='solid', fgColor='FFFF00')
sheet['A2'].fill = fill
sheet['A2'] = 'Yellow Background'
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
sheet['A3'].border = thin_border
sheet['A3'] = 'Bordered Cell'
align = Alignment(horizontal='center', vertical='center')
sheet['A4'].alignment = align
sheet['A4'] = 'Center Aligned'
# 保存工作簿
wb.save('styled.xlsx')
import qstock as qs

market_data = qs.get_market_data(date="2023-03-30") # 示例日期，调整为需要的日期

# 假设 market_data 包含了所有需要的市场数据信息
# 构建一个字典，后续转换为 pandas DataFrame
data = {
    "日期": ["2023-03-30"],
    "A股公司数量": [market_data["a_shares_count"]],
    "深圳交易所交易额": [market_data["sz_exchange_volume"]],
    "上海交易所交易额": [market_data["sh_exchange_volume"]],
    "涨幅超过10%的公司数量": [market_data["up_more_than_10_percent_count"]],
    # 依此类推，填充其他数据
}

# 将字典转换为 DataFrame
df = pd.DataFrame(data)

# 输出到Excel文件
df.to_excel("market_data.xlsx", index=False)
